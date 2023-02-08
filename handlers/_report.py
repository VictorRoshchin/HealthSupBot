from calendar import monthrange
from contextlib import suppress
from datetime import datetime, date
from time import strptime

import openpyxl
from aiogram import types, Dispatcher
from aiogram.dispatcher.filters import Text
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, InputFile
from aiogram.utils.callback_data import CallbackData
from aiogram.utils.exceptions import MessageNotModified

from db_api import api_requests
from create_bot import bot
from handlers._report_mail import send_report
from keyboards.all import kb_main_menu

period_report = CallbackData('period_report', 'action', 'year', 'b_year', 'month', 'b_month', 'day', 'b_day', 'select_from')


async def report_create_file(user_id, entries: list):
    etrs = [list(e)[:-1] for e in entries]
    header = ['Верхнее давление', 'Нижнее давление', 'Пульс', 'Дата', 'Время']
    report = openpyxl.Workbook()
    report.create_sheet(title='Отчет', index=0)
    table = report['Отчет']
    for row in range(1, len(etrs) + 2):
        if row == 1:
            for col in range(1, len(header)+1):
                value = header[col-1]
                cell = table.cell(row=row, column=col)
                cell.value = value
        else:
            for col in range(1, len(etrs[row-2])+1):
                value = etrs[row-2][col-1]
                cell = table.cell(row=row, column=col)
                cell.value = value
    report.save(f'report_{user_id}.xlsx')
    return f'report_{user_id}.xlsx'


# Формирование и отправка отчета.
async def report(message: types.Message):
    await message.reply('Сформатировать полный отчет или выбрать период?',
                        reply_markup=InlineKeyboardMarkup().add(
                            InlineKeyboardButton(text='Полный отчет', callback_data='report_full'),
                            InlineKeyboardButton(text='Выбрать даты', callback_data=period_report.new(
                                action='get_year',
                                year='',
                                b_year='',
                                month='',
                                b_month='',
                                day='',
                                b_day='',
                                select_from='',
                            )),
                        ))


async def report_to_telegram_or_email(report_to, call, report_file):
    if report_to == 'telegram':
        await call.message.answer('Отчет успешно сформирован и отправлен',
                                  reply_markup=kb_main_menu())
        await call.message.answer_document(InputFile(report_file))

    else:
        await send_report(report_to, report_file)
        await call.message.answer(f'Отчет отправлен на электронную почту {report_to}',
                                  reply_markup=kb_main_menu())


async def report_full(call: types.CallbackQuery):
    user_id = call.message.reply_to_message.from_user.id
    await call.answer(text='Формируем отчет')
    await call.message.delete()
    entries_all = await api_requests.get(user_id)
    report_to = await api_requests.get(user_id, report_to=True)
    report_file = await report_create_file(user_id, entries_all)

    await report_to_telegram_or_email(report_to, call, report_file)


async def keyboard_years(years, b_year, b_month, b_day, select_from, callback, action):
    kb = InlineKeyboardMarkup()
    for y in years:
        kb.insert(
            InlineKeyboardButton(text=str(y), callback_data=callback.new(
                action=action,
                year=y, b_year=b_year,
                month=b_month, b_month=b_month,
                day=b_day, b_day=b_day,
                select_from=select_from
            ))
        )
    return kb


async def report_period_year(call: types.CallbackQuery, callback_data: dict):
    user_id = call.from_user.id
    await call.message.delete()
    await call.answer()
    select_from = callback_data.get('select_from', '')

    b_date = await api_requests.get(user_id, first_entry_date=True)

    b_year, b_month, b_day = map(int, str(b_date).split('-'))

    t_year, t_month, t_day = map(int, datetime.today().date().strftime("%Y %m %d").split())

    b_year = int(b_year)

    years = (y for y in range(b_year, int(t_year) + 1))

    kb = await keyboard_years(years, b_year, b_month, b_day, select_from, period_report, 'get_month')

    await bot.send_message(user_id, 'Выберите год', reply_markup=kb)


async def keyboard_months(months, year, b_year, b_month, b_day, select_from, callback, action):
    kb = InlineKeyboardMarkup()
    for m in months:
        kb.insert(
            InlineKeyboardButton(text=str(m), callback_data=callback.new(
                action=action,
                year=year, b_year=b_year,
                month=m, b_month=b_month,
                day=b_day, b_day=b_day,
                select_from=select_from
            ))
        )
    return kb


async def report_period_month(call: types.CallbackQuery, callback_data: dict):
    user_id = call.from_user.id
    await call.message.delete()
    await call.answer()
    select_from = callback_data.get('select_from', '')
    year = int(callback_data['year'])
    b_year = int(callback_data['b_year'])
    b_month = int(callback_data['b_month'])
    b_day = int(callback_data['b_day'])

    t_year, t_month, t_day = map(int, datetime.today().date().strftime("%Y %m %d").split())

    if year == b_year and year == t_year:
        months = (y for y in range(b_month, t_month + 1))
    elif year == b_year:
        months = (y for y in range(b_month, 13))
    elif year == t_year:
        months = (y for y in range(1, t_month + 1))
    else:
        months = (y for y in range(1, 13))

    kb = await keyboard_months(months, year, b_year, b_month, b_day, select_from, period_report, 'get_day')

    await bot.send_message(user_id, 'Выберите месяц', reply_markup=kb)


async def keyboard_days(days, year, b_year, month, b_month, b_day, select_from, callback, action):
    kb = InlineKeyboardMarkup()
    for d in days:
        kb.insert(
            InlineKeyboardButton(text=str(d), callback_data=callback.new(
                action=action,
                year=year, b_year=b_year,
                month=month, b_month=b_month,
                day=d, b_day=b_day,
                select_from=select_from
            ))
        )
    return kb


async def report_period_day(call: types.CallbackQuery, callback_data: dict):
    user_id = call.from_user.id
    await call.message.delete()
    await call.answer()
    select_from = callback_data.get('select_from', '')
    year = int(callback_data['year'])
    month = int(callback_data['month'])
    b_year = int(callback_data['b_year'])
    b_month = int(callback_data['b_month'])
    b_day = int(callback_data['b_day'])

    weekday, last_day = map(int, monthrange(year, month))

    t_year, t_month, t_day = map(int, datetime.today().date().strftime("%Y %m %d").split())

    if month == b_month and month == t_month:
        days = range(b_day, t_day + 1)

    elif month == b_month:
        days = range(b_day, last_day + 1)

    elif month == t_month:
        days = range(1, t_day + 1)

    else:
        days = range(1, last_day+1)

    kb = await keyboard_days(days, year, b_year, month, b_month,
                             b_day, select_from, period_report, 'accept_date_from')
    await bot.send_message(user_id, 'Выберите число', reply_markup=kb)


accept_from = CallbackData('p_from', 'action', 'date_append', 'select_from')


async def date_f(call: types.CallbackQuery, callback_data: dict):
    user_id = call.from_user.id
    await call.message.delete()
    await call.answer()
    select_from = callback_data.get('select_from', '')
    day = callback_data['day']
    year = callback_data['year']
    month = callback_data['month']
    mth = [0, 'январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
           'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
    await bot.send_message(user_id, f'Выбран {year} год, {month} месяц ({mth[int(month)]}), {day} число',
                           reply_markup=InlineKeyboardMarkup()
                           .add(InlineKeyboardButton(text='Подтвердить',
                                                     callback_data=accept_from.new(
                                                         action='append_date_from',
                                                         date_append=f'{year}-{month}-{day}',
                                                         select_from=select_from
                                                     ))
                                )
                           .add(
                               InlineKeyboardButton(text='Изменить дату',
                                                    callback_data=period_report.new(
                                                        action='get_year',
                                                        year='',
                                                        b_year='',
                                                        month='',
                                                        b_month='',
                                                        day='',
                                                        b_day='',
                                                        select_from=select_from,

                                                    ))
                           ))


append_date = CallbackData('accept_date', 'action', 'date_to', 'date_from')


async def period_to(call: types.CallbackQuery, callback_data: dict):
    user_id = call.from_user.id
    await call.message.delete()
    await call.answer()

    yearf, monthf, dayf = callback_data["date_append"].split('-')
    date_append = f'{yearf}-{monthf.rjust(2, "0")}-{dayf.rjust(2, "0")}'

    select_from = callback_data.get('select_from', '')

    if select_from != '':
        yeart, montht, dayt = select_from.split('-')
        select_from = f'{yeart}-{montht.rjust(2, "0")}-{dayt.rjust(2, "0")}'
        await call.message.answer(f'Отчет будет сформирован за период:\n'
                                  f'От {select_from} до {date_append} включительно',
                                  reply_markup=InlineKeyboardMarkup().add(
                                      InlineKeyboardButton(text='Подтвердить',
                                                           callback_data=append_date.new(
                                                               action='custom_report',
                                                               date_to=date_append,
                                                               date_from=select_from
                                                           )),
                                      InlineKeyboardButton(text='Изменить', callback_data=period_report.new(
                                          action='get_year',
                                          year='',
                                          b_year='',
                                          month='',
                                          b_month='',
                                          day='',
                                          b_day='',
                                          select_from=select_from,
                                      ))
                                  ))
    else:
        await bot.send_message(user_id,
                               f'Выбрана начальная дата {date_append}.\n'
                               f'До какой даты включительно формировать отчет?',
                               reply_markup=InlineKeyboardMarkup().add(
                                   InlineKeyboardButton(text='До настоящего времени',
                                                        callback_data=accept_from.new(
                                                            action='current_report',
                                                            date_append=date_append,
                                                            select_from=''
                                                         )),
                                   InlineKeyboardButton(text='Выбрать дату', callback_data=period_report.new(
                                       action='get_year',
                                       year='',
                                       b_year='',
                                       month='',
                                       b_month='',
                                       day='',
                                       b_day='',
                                       select_from=date_append,
                                   ))
                               ))


async def report_to_current(call: types.CallbackQuery, callback_data: dict):
    user_id = call.from_user.id
    await call.answer(text='Формируем отчет')
    await call.message.delete()

    dt_str = callback_data["date_append"]

    entries = await api_requests.get(user_id, date_from=dt_str)
    report_to = await api_requests.get(user_id, report_to=True)
    report_file = await report_create_file(user_id, entries)

    await report_to_telegram_or_email(report_to, call, report_file)


async def report_to_custom(call: types.CallbackQuery, callback_data: dict):
    user_id = call.from_user.id
    await call.answer(text='Формируем отчет')
    await call.message.delete()
    df = callback_data["date_from"]
    dt = callback_data["date_to"]
    entries = await api_requests.get(user_id, date_from=df, date_to=dt)
    report_to = await api_requests.get(user_id, report_to=True)
    report_file = await report_create_file(user_id, entries)

    await report_to_telegram_or_email(report_to, call, report_file)


def register_handlers_report(disp: Dispatcher):
    disp.register_message_handler(report, Text(equals='Отчет', ignore_case=True))

    # Полного отчета.
    disp.register_callback_query_handler(report_full, lambda x: x.data and x.data.startswith('report_full'))

    # Выбор периода.
    # Дата ОТ.
    disp.register_callback_query_handler(report_period_year, period_report.filter(action=['get_year']))
    disp.register_callback_query_handler(report_period_month, period_report.filter(action=['get_month']))
    disp.register_callback_query_handler(report_period_day, period_report.filter(action=['get_day']))
    # Подтверждение даты ОТ.
    disp.register_callback_query_handler(date_f, period_report.filter(action=['accept_date_from']))
    # Выбор даты ДО.
    disp.register_callback_query_handler(period_to, accept_from.filter(action=['append_date_from']))

    disp.register_callback_query_handler(report_to_current, accept_from.filter(action=['current_report']))
    disp.register_callback_query_handler(report_to_custom, append_date.filter(action=['custom_report']))
